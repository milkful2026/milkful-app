"""Create Flutter mobile app stories from NSMB Excel under epic MA-18."""
import base64
import json
import os
import re
import urllib.error
import urllib.request
from openpyxl import load_workbook

EMAIL = os.environ.get("JIRA_EMAIL", "milkfuldairyindia@gmail.com")
TOKEN = os.environ["JIRA_TOKEN"]
BASE = "https://milkfuldairyindia.atlassian.net"
EPIC_KEY = "MA-18"
EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "docs", "jira", "attachments", "NSMB.xlsx"
)

AUTH = base64.b64encode(f"{EMAIL}:{TOKEN}".encode()).decode()
HEADERS = {
    "Authorization": f"Basic {AUTH}",
    "Accept": "application/json",
    "Content-Type": "application/json",
}

MOBILE_FEATURES = {
    "User Registration",
    "User Login",
    "Product Listing (Inventory-driven)",
    "Add to Cart",
    "Payment Gateway Integration",
    "Subscription Module (Start / Stop / Pause)",
    "Order History",
    "Transaction History",
    "Feedback (Delivery / Product Quality)",
    "Offers",
    "Referrals",
    "Calendar View (Date-based Ordering)",
    "Order Cancellation",
    "Wallet Recharge",
    "Order Confirmation / Preview",
    "Order Details",
    "Order Status",
    "Inventory Module Integration",
    "Invoice Generation (GST, Discounts, Offers)",
}

SKIP_EXISTING = {"User Registration": "MA-1"}


def api(method, url, data=None):
    body = json.dumps(data).encode() if data is not None else None
    req = urllib.request.Request(url, data=body, headers=HEADERS, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            raw = resp.read().decode()
            return json.loads(raw) if raw.strip() else {}
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"{method} {url} -> {e.code}: {e.read().decode()}") from e


def text(value, marks=None):
    node = {"type": "text", "text": value}
    if marks:
        node["marks"] = marks
    return node


def heading(level, value):
    return {"type": "heading", "attrs": {"level": level}, "content": [text(value)]}


def paragraph(*parts):
    return {"type": "paragraph", "content": list(parts)}


def bullet(items):
    return {
        "type": "bulletList",
        "content": [
            {"type": "listItem", "content": [paragraph(text(item))]} for item in items
        ],
    }


def build_description(feature):
    name = feature["name"]
    module = feature["module"]
    reqs = feature["requirements"]
    deps = feature["dependencies"]
    priority = feature["priority"]

    return {
        "version": 1,
        "type": "doc",
        "content": [
            heading(2, "User Story"),
            paragraph(
                text("As a ",),
                text("Milkful customer", [{"type": "strong"}]),
                text(f", I want {name.lower()} functionality in the mobile app, so that I can use the full milk delivery service on Android and iOS."),
            ),
            heading(2, "Module"),
            paragraph(text(module)),
            heading(2, "Platform"),
            paragraph(text("Flutter (Android and iOS)")),
            heading(2, "Functional Requirements"),
            paragraph(text(reqs)),
            heading(2, "Key Dependencies"),
            paragraph(text(deps)),
            heading(2, "Priority"),
            paragraph(text(priority)),
            heading(2, "Source"),
            paragraph(text("NSMB App Feature Spec (attached to epic MA-18)")),
        ],
    }


def priority_label(priority):
    p = (priority or "").strip().lower()
    if "must" in p:
        return "must-have"
    if "should" in p:
        return "should-have"
    if "could" in p:
        return "could-have"
    return "mobile"


def read_mobile_features():
    path = os.path.normpath(EXCEL_PATH)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["App Feature Spec"]
    features = []
    in_admin = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        joined = " ".join(cells).upper()
        if "ADMIN / BACK-OFFICE" in joined:
            in_admin = True
            continue
        if in_admin:
            continue

        num = cells[0] if cells else ""
        if not re.match(r"^\d+(\.0)?$", num):
            continue

        name = cells[1] if len(cells) > 1 else ""
        if name not in MOBILE_FEATURES:
            continue

        features.append(
            {
                "num": num.replace(".0", ""),
                "name": name,
                "module": cells[2] if len(cells) > 2 else "",
                "requirements": cells[3] if len(cells) > 3 else "",
                "dependencies": cells[4] if len(cells) > 4 else "",
                "priority": cells[5] if len(cells) > 5 else "",
            }
        )

    wb.close()
    return features


def link_existing(issue_key):
    api(
        "PUT",
        f"{BASE}/rest/api/3/issue/{issue_key}",
        {"fields": {"parent": {"key": EPIC_KEY}}},
    )


def create_story(feature):
    summary = f"{feature['name']} — {feature['module']} (Flutter)"
    labels = ["flutter", "mobile", "nsmb", priority_label(feature["priority"])]
    body = {
        "fields": {
            "project": {"key": "MA"},
            "parent": {"key": EPIC_KEY},
            "summary": summary,
            "issuetype": {"id": "10003"},
            "description": build_description(feature),
            "labels": labels,
        }
    }
    return api("POST", f"{BASE}/rest/api/3/issue", body)


def main():
    features = read_mobile_features()
    print(f"Found {len(features)} mobile features in NSMB Excel")

    for name, key in SKIP_EXISTING.items():
        link_existing(key)
        print(f"Linked existing {key} ({name}) -> {EPIC_KEY}")

    created = []
    skipped = []
    for feature in features:
        if feature["name"] in SKIP_EXISTING:
            skipped.append(feature["name"])
            continue
        result = create_story(feature)
        created.append((result["key"], feature["name"]))
        print(f"{result['key']} - {feature['num']}. {feature['name']}")

    print(f"\nCreated {len(created)} stories, skipped {len(skipped)} existing")
    print(f"Epic: {BASE}/browse/{EPIC_KEY}")


if __name__ == "__main__":
    main()
