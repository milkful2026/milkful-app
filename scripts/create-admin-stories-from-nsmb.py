"""Create Admin-UI stories from NSMB Excel under epic MA-20."""
import base64
import json
import os
import re
import urllib.error
import urllib.request
from openpyxl import load_workbook

EMAIL = os.environ.get("JIRA_EMAIL", "milkfuldairyindia@gmail.com")
TOKEN = os.environ["JIRA_TOKEN"]
BASE = "https://milkfuldairyindia.atlassian.net"
EPIC_KEY = "MA-20"
EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "docs", "jira", "attachments", "NSMB.xlsx"
)

AUTH = base64.b64encode(f"{EMAIL}:{TOKEN}".encode()).decode()
HEADERS = {
    "Authorization": f"Basic {AUTH}",
    "Accept": "application/json",
    "Content-Type": "application/json",
}


def api(method, url, data=None):
    body = json.dumps(data).encode() if data is not None else None
    req = urllib.request.Request(url, data=body, headers=HEADERS, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            raw = resp.read().decode()
            return json.loads(raw) if raw.strip() else {}
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"{method} {url} -> {e.code}: {e.read().decode()}") from e


def text(value, marks=None):
    node = {"type": "text", "text": value}
    if marks:
        node["marks"] = marks
    return node


def heading(level, value):
    return {"type": "heading", "attrs": {"level": level}, "content": [text(value)]}


def paragraph(*parts):
    return {"type": "paragraph", "content": list(parts)}


def build_description(feature):
    return {
        "version": 1,
        "type": "doc",
        "content": [
            heading(2, "User Story"),
            paragraph(
                text("As an ",),
                text("admin/operations user", [{"type": "strong"}]),
                text(
                    f", I want {feature['name'].lower()} in the admin console, "
                    "so that I can manage the milk delivery business effectively."
                ),
            ),
            heading(2, "Module"),
            paragraph(text(feature["module"])),
            heading(2, "Platform"),
            paragraph(text("Admin Web UI")),
            heading(2, "Functional Requirements"),
            paragraph(text(feature["requirements"])),
            heading(2, "Key Dependencies"),
            paragraph(text(feature["dependencies"])),
            heading(2, "Priority"),
            paragraph(text(feature["priority"])),
            heading(2, "Source"),
            paragraph(text("NSMB App Feature Spec — Admin / Back-Office section")),
        ],
    }


def priority_label(priority):
    p = (priority or "").strip().lower()
    if "must" in p:
        return "must-have"
    if "should" in p:
        return "should-have"
    if "could" in p:
        return "could-have"
    return "admin"


def read_admin_features():
    path = os.path.normpath(EXCEL_PATH)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["App Feature Spec"]
    features = []
    in_admin = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        joined = " ".join(cells).upper()
        if "ADMIN / BACK-OFFICE" in joined:
            in_admin = True
            continue
        if not in_admin:
            continue

        num = cells[0] if cells else ""
        if not re.match(r"^\d+(\.0)?$", num):
            continue

        features.append(
            {
                "num": num.replace(".0", ""),
                "name": cells[1] if len(cells) > 1 else "",
                "module": cells[2] if len(cells) > 2 else "",
                "requirements": cells[3] if len(cells) > 3 else "",
                "dependencies": cells[4] if len(cells) > 4 else "",
                "priority": cells[5] if len(cells) > 5 else "",
            }
        )

    wb.close()
    return features


def create_story(feature):
    summary = f"{feature['name']} — {feature['module']} (Admin UI)"
    labels = ["admin-ui", "nsmb", priority_label(feature["priority"])]
    body = {
        "fields": {
            "project": {"key": "MA"},
            "parent": {"key": EPIC_KEY},
            "summary": summary,
            "issuetype": {"id": "10003"},
            "description": build_description(feature),
            "labels": labels,
        }
    }
    return api("POST", f"{BASE}/rest/api/3/issue", body)


def main():
    features = read_admin_features()
    print(f"Found {len(features)} admin features in NSMB Excel")

    created = []
    for feature in features:
        result = create_story(feature)
        created.append((result["key"], feature["name"]))
        print(f"{result['key']} - {feature['num']}. {feature['name']}")

    print(f"\nCreated {len(created)} stories under {EPIC_KEY}")
    print(f"Epic: {BASE}/browse/{EPIC_KEY}")


if __name__ == "__main__":
    main()
